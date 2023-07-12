import openpyxl


def find_texts_in_column(file_path, sheet_name, column_index, search_texts):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    print(sheet)

    found_texts = []

    for search_text in search_texts:
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=column_index, max_col=column_index):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value == search_text:
                    found_texts.append(search_text)
                    break

    return found_texts

# Example usage
file_path = 'filename.xlsx'  # Replace with your Excel file path
sheet_name = 'sheet1'  # Replace with your sheet name
column_index = 2  # Replace with the column index (e.g., 1 for column A, 2 for column B, etc.)
search_texts = ['element1', 'element2', 'element3', 'element4']
found_texts = find_texts_in_column(file_path, sheet_name, column_index, search_texts)
if found_texts:
    print("The following texts were found in the specified column:")
    for text in found_texts:
        print(text)
else:
    print("None of the search texts were found in the specified column.")
